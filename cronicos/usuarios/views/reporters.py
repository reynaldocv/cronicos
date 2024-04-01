from django.shortcuts import redirect, render
from django.template import loader
from django.http import HttpResponse
from django.db.models import Q
from re import template

from datetime import date

from usuarios.views.reporters import *
import os

from openpyxl import Workbook
from openpyxl.styles import Font
import openpyxl 

from usuarios.modules import codeHtml

from usuarios.models import Atencion
from usuarios.models import Paciente
from usuarios.models import Hospital
from usuarios.models import Especialidad
from usuarios.models import Pie
from usuarios.models import Ekg
from usuarios.models import Referencia



meses = ["", "Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Set", "Oct", "Nov", "Dic"]
Meses = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Setiembre", "Octubre", "Noviembre", "Diciembre"]


def reporte_01(request):
    _Mhta_count = len(Paciente.objects.filter(muerto = 0, hta = True, dm = False, sexo = True))
    _Fhta_count = len(Paciente.objects.filter(muerto = 0, hta = True, dm = False, sexo = False))
    _Mdm_count = len(Paciente.objects.filter(muerto = 0, hta = False, dm = True, sexo = True))
    _Fdm_count = len(Paciente.objects.filter(muerto = 0, hta = False, dm = True, sexo = False))
    _MhtaDm_count = len(Paciente.objects.filter(muerto = 0, hta = True, dm = True, sexo = True))
    _FhtaDm_count = len(Paciente.objects.filter(muerto = 0, hta = True, dm = True, sexo = False))

    _total = len(Paciente.objects.filter(muerto = 0))
    _totalSum = _Mhta_count + _Fhta_count + _Mdm_count + _Fdm_count +  _MhtaDm_count +  _FhtaDm_count

    _context = {"usuarios": {"hombres": {"dm": _Mdm_count, "hta": _Mhta_count, "htaDm": _MhtaDm_count}, 
                            "mujeres": {"dm": _Fdm_count, "hta": _Fhta_count, "htaDm": _FhtaDm_count}, }, 
               "total" : {"hta": _Mhta_count + _Fhta_count, 
                          "dm" : _Mdm_count + _Fdm_count, 
                          "htaDm": _MhtaDm_count + _FhtaDm_count,
                          "suma": _totalSum, "database": _total}, 
               }

    _template = loader.get_template("reporte_01.html")

    return HttpResponse(_template.render(_context, request))

def reporte_02(request):
    _year = date.today().year

    if "nowYear" in request.POST:
        _year = request.POST["nowYear"]

    _context = {}
    _total = len(Pie.objects.filter(data__year = _year))
       
    _context["DmHta"] = []
    _context["Dm"] = []
    _context["Hta"] = []
    _context["totalMes"] = []

    _totalDmHta = 0
    _totalDm = 0
    _totalHta = 0

    for mes in range(1, 13):
        _Dm = Paciente.objects.filter(dm = True, fechaDm__month = mes, fechaDm__year = _year)
        _Hta = Paciente.objects.filter(hta = True, fechaHta__month = mes, fechaHta__year = _year)

        counter = [0, 0]

        for paciente in _Dm: 
            if paciente.hta == False: 
                counter[0] += 1 
            
            elif paciente.fechaHta.month == paciente.fechaDm.month and  paciente.fechaHta.year == paciente.fechaDm.year:
                counter[0] += 1 

            elif paciente.fechaDm < paciente.fechaHta: 
                counter[0] += 1 
        
        for paciente in _Hta:             
            if paciente.dm == False: 
                counter[1] += 1 
            
            elif paciente.fechaHta.month == paciente.fechaDm.month and  paciente.fechaHta.year == paciente.fechaDm.year:
                counter[1] += 1 

            elif paciente.fechaDm > paciente.fechaHta: 
                counter[1] += 1 

        _totalDm += counter[0]
        _totalHta += counter[1]
        
        _context["Dm"].append(counter[0]) 
        _context["Hta"].append(counter[1]) 
        _context["totalMes"].append(sum(counter))
    
    
    _context["Dm"].append(_totalDm) 
    _context["Hta"].append(_totalHta) 

    _context["totalMes"].append(_totalDm + _totalHta)
    _context["years"] = {"now": _year}
    _context["meses"] = [meses[i] for i in range(1, 13)] + ["total"]

    _template = loader.get_template("reporte_02.html")

    return HttpResponse(_template.render(_context, request))

def reporte_pie(request):
    _year = date.today().year

    if "nowYear" in request.POST:
        _year = request.POST["nowYear"]

    _context = {}
    _total = len(Pie.objects.filter(data__year = _year))
       
    _context["conLesion"] = []
    _context["sinLesion"] = []
    _context["totalmes"] = []

    _totalconLesion = 0
    _totalsinLesion = 0

    for mes in range(1, 13):
        _conLesion = len(Pie.objects.filter(data__month = mes, data__year = _year, conLesion = True ))
        _sinLesion = len(Pie.objects.filter(data__month = mes, data__year = _year, conLesion = False ))

        _totalconLesion += _conLesion
        _totalsinLesion += _sinLesion

        _context["conLesion"].append(_conLesion) 
        _context["sinLesion"].append(_sinLesion) 
        _context["totalmes"].append(_conLesion + _sinLesion) 
    
    _context["conLesion"].append(_totalconLesion)
    _context["sinLesion"].append(_totalsinLesion)

    _context["total"] = _total
    _context["totalmes"].append(_totalconLesion + _totalsinLesion) 
    _context["years"] = {"now": _year}
    _context["meses"] = [meses[i] for i in range(1, 13)] + ["total"]

    _template = loader.get_template("reporte_pie.html")

    return HttpResponse(_template.render(_context, request))


def reporte_ekg(request):
    _year = date.today().year

    if "nowYear" in request.POST:
        _year = request.POST["nowYear"]

    _context = {}
    _total = len(Ekg.objects.filter(data__year = _year))
       
    _context["normal"] = []
    _context["alterado"] = []
    _context["totalmes"] = []

    _totalNormal = 0
    _totalAlterado = 0

    for mes in range(1, 13):
        _normal = len(Ekg.objects.filter(data__month = mes, data__year = _year, esNormal = True ))
        _alterado = len(Ekg.objects.filter(data__month = mes, data__year = _year, esNormal = False ))

        _totalNormal += _normal
        _totalAlterado += _alterado

        _context["normal"].append(_normal) 
        _context["alterado"].append(_alterado) 
        _context["totalmes"].append(_normal + _alterado) 
    
    _context["normal"].append(_totalNormal)
    _context["alterado"].append(_totalAlterado)

    _context["total"] = _total
    _context["totalmes"].append(_totalNormal + _totalAlterado) 
    _context["years"] = {"now": _year}
    _context["meses"] = [meses[i] for i in range(1, 13)] + ["total"]

    _template = loader.get_template("reporte_ekg.html")

    return HttpResponse(_template.render(_context, request))



def reporte_dot(request):
    _year = date.today().year

    if "nowYear" in request.POST:
        _year = request.POST["nowYear"]

    _context = {}
    _total = len(Atencion.objects.filter(fecha__year = _year))
       
    _context["dotacion1"] = []
    _context["dotacion2"] = []
    _context["dotacion3"] = []
    _context["dotacion0"] = []
    _context["totalmes"] = []

    _totalDotacion = [0, 0, 0, 0]
    
    for mes in range(1, 13):
        _dot0 = len(Atencion.objects.filter(fecha__month = mes, fecha__year = _year, dotacion = 0 ))
        _dot1 = len(Atencion.objects.filter(fecha__month = mes, fecha__year = _year, dotacion = 1 ))
        _dot2 = len(Atencion.objects.filter(fecha__month = mes, fecha__year = _year, dotacion = 2 ))
        _dot3 = len(Atencion.objects.filter(fecha__month = mes, fecha__year = _year, dotacion = 3 ))

        _totalDotacion[0] += _dot0
        _totalDotacion[1] += _dot1
        _totalDotacion[2] += _dot2        
        _totalDotacion[3] += _dot3
        
        _context["dotacion0"].append(_dot0) 
        _context["dotacion1"].append(_dot1) 
        _context["dotacion2"].append(_dot2) 
        _context["dotacion3"].append(_dot3) 

        _context["totalmes"].append(_dot0 + _dot1 + _dot2 + _dot3)
     
    _context["dotacion0"].append(_totalDotacion[0]) 
    _context["dotacion1"].append(_totalDotacion[1])
    _context["dotacion2"].append(_totalDotacion[2])
    _context["dotacion3"].append(_totalDotacion[3])

    _context["total"] = _total
    _context["totalmes"].append(sum(_totalDotacion)) 
    _context["years"] = {"now": _year}
    _context["meses"] = [meses[i] for i in range(1, 13)] + ["total"]

    _template = loader.get_template("reporte_dot.html")

    return HttpResponse(_template.render(_context, request))


def reporte_ref(request):
    _year = date.today().year

    if "nowYear" in request.POST:
        _year = request.POST["nowYear"]

    _context = {}
    _total = len(Referencia.objects.filter(data__year = _year))
    _totalA = len(Referencia.objects.filter(data__year = _year, atendido = True))

    _especialidades = Especialidad.objects.all()
    _totalEsp = {}
    
    _context["estadisticas"] = {}
        
    for (i, _esp) in enumerate(_especialidades): 
        _context["estadisticas"][_esp.id] = [_esp.especialidad]

    _context["totalmes"] = []

    _refT = {}    
    _refA = {}

    for mes in range(1, 13):
        
        _refTotal = 0
        _refAtten = 0 

        for (i, _esp) in enumerate(_especialidades):                         
            refTotal = len(Referencia.objects.filter(data__year = _year, data__month = mes, especialidad = _esp)) 
            refAtten = len(Referencia.objects.filter(data__year = _year, data__month = mes, especialidad = _esp, atendido = True)) 

            _context["estadisticas"][_esp.id].append(str(refTotal) + " (" + str(refAtten) + ")")

            _refT[_esp.id] = _refT.get(_esp.id, 0) + refTotal
            _refA[_esp.id] = _refA.get(_esp.id, 0) + refAtten

            _refTotal += refTotal
            _refAtten += refAtten

        _context["totalmes"].append(str(_refTotal) + " (" + str(_refAtten) + ")") 

    _refTotal = 0
    _refAtten = 0 

    for (i, _esp) in enumerate(_especialidades): 
        _context["estadisticas"][_esp.id].append(str(_refT[_esp.id]) +  " (" + str(_refA[_esp.id]) + ")")
        
        _refTotal += _refT[_esp.id]
        _refAtten += _refA[_esp.id]

    _context["totalmes"].append(str(_refTotal) + " (" + str(_refAtten) + ")") 

    _context["total"] = str(_total) + " (" + str(_totalA) + ")"
    
    _context["years"] = {"now": _year}
    _context["meses"] = [meses[i] for i in range(1, 13)] + ["total"]
    
    _template = loader.get_template("reporte_ref.html")
    
    return HttpResponse(_template.render(_context, request))

def reporte_imc(request):
    _year = date.today().year

    if "nowYear" in request.POST:
        _year = request.POST["nowYear"]

    categorias = ["Normal", "Sobrepeso", "Obeso"]

    _context = {}
    _total = len(Atencion.objects.filter(fecha__year = _year, imcDescripcion="Obeso"))
    _total += len(Atencion.objects.filter(fecha__year = _year, imcDescripcion="Sobrepeso"))
    _total += len(Atencion.objects.filter(fecha__year = _year, imcDescripcion="Normal"))
    
    _totalEsp = {}
    
    _context["estadisticas"] = {}
        
    for (i, _esp) in enumerate(categorias): 
        _context["estadisticas"][str(i)] = [_esp]

    _context["totalmes"] = []

    _refT = {}    
    _refA = {}
    for mes in range(1, 13):
        
        _refTotal = 0
        _refAtten = 0 

        for (i, _esp) in enumerate(categorias):                         
            refTotal = len(Atencion.objects.filter(fecha__year = _year, fecha__month = mes, imcDescripcion = _esp)) 
            #refTotal = 0
            
            _context["estadisticas"][str(i)].append(refTotal)

            _refT[str(i)] = _refT.get(str(i), 0) + refTotal            

            _refTotal += refTotal
            
        _context["totalmes"].append(_refTotal) 

    _refTotal = 0
    _refAtten = 0 

    for (i, _esp) in enumerate(categorias): 
        _context["estadisticas"][str(i)].append(_refT[str(i)])
        
        _refTotal += _refT[str(i)]
    

    _context["totalmes"].append(_refTotal) 
    _context["total"] = _total
    _context["years"] = {"now": _year}
    _context["meses"] = [meses[i] for i in range(1, 13)] + ["total"]
    
    _template = loader.get_template("reporte_imc.html")
    
    return HttpResponse(_template.render(_context, request))

def reporte_deadman(request):
    _year = date.today().year

    if "nowYear" in request.POST:
        _year = request.POST["nowYear"]

    _context = {}
    _total = len(Paciente.objects.filter(muerto = True, fechaMuerto__year = _year))
       
    _context["DmHta"] = []
    _context["Dm"] = []
    _context["Hta"] = []
    
    _context["totalmes"] = []

    _totalDotacion = [0, 0, 0, 0]
    
    for mes in range(1, 13):
        _dot0 = len(Paciente.objects.filter(muerto = True, fechaMuerto__month = mes, fechaMuerto__year = _year, hta = True, dm = True ))
        _dot1 = len(Paciente.objects.filter(muerto = True, fechaMuerto__month = mes, fechaMuerto__year = _year, hta = False, dm = True ))
        _dot2 = len(Paciente.objects.filter(muerto = True, fechaMuerto__month = mes, fechaMuerto__year = _year, hta = True, dm = False ))
        
        _totalDotacion[0] += _dot0
        _totalDotacion[1] += _dot1
        _totalDotacion[2] += _dot2        
        
        _context["DmHta"].append(_dot0) 
        _context["Dm"].append(_dot1) 
        _context["Hta"].append(_dot2) 
        
        _context["totalmes"].append(_dot0 + _dot1 + _dot2)
     
    _context["DmHta"].append(_totalDotacion[0]) 
    _context["Dm"].append(_totalDotacion[1])
    _context["Hta"].append(_totalDotacion[2])
    
    _context["total"] = _total
    _context["totalmes"].append(sum(_totalDotacion)) 
    _context["years"] = {"now": _year}
    _context["meses"] = [meses[i] for i in range(1, 13)] + ["total"]

    _template = loader.get_template("reporte_deadman.html")

    return HttpResponse(_template.render(_context, request))






    

    

    
    



    































def excel_01(request):
    tmp = date.today()

    hipertensos = Paciente.objects.filter(muerto = False, hta = True, dm = False).order_by("nombre")
    deabeticos = Paciente.objects.filter(muerto = False, hta = False, dm = True).order_by("nombre")
    ambos = Paciente.objects.filter(muerto = False, hta = True, dm = True).order_by("nombre")

    muertos = Paciente.objects.filter(muerto = True); 

    red_font = Font(color = '00FF0000', bold = True)
    blue_font = Font(color = '00009900', bold = True, size=12)

    wb = Workbook()

    sheet = wb.active
    sheet.title = "Total"

    sheet.cell(row=4, column=2).font = red_font
    sheet.cell(row=4, column=2).value = "Pacientes con HTAxDM"
    sheet.cell(row=4, column=3).value = len(ambos)
    
    sheet.cell(row=5, column=2).font = red_font
    sheet.cell(row=5, column=2).value = "Pacientes con DM"    
    sheet.cell(row=5, column=3).value = len(deabeticos)

    sheet.cell(row=6, column=2).font = red_font
    sheet.cell(row=6, column=2).value = "Pacientes con HTA"
    sheet.cell(row=6, column=3).value = len(hipertensos)

    sheet.cell(row=7, column=2).font = red_font
    sheet.cell(row=7, column=2).value = "Total"
    sheet.cell(row=7, column=3).value = len(hipertensos) + len(deabeticos) + len(ambos)

    sheet.cell(row=9, column=2).font = red_font
    sheet.cell(row=9, column=2).value = "Paciente muertos"
    sheet.cell(row=9, column=3).value = len(muertos)

    sheet.cell(row=4, column=5).value = date.today()
    
    sheet = wb.create_sheet(title = "HTA")

    sheet.cell(2, 2).value = "PACIENTES CON HTA"
    sheet.cell(2, 2).font = blue_font
    

    sheet.cell(4, 1).value = "#"
    sheet.cell(4, 2).value = "NOMBRE"
    sheet.cell(4, 3).value = "NACIMIENTO"
    sheet.cell(4, 4).value = "EDAD"
    sheet.cell(4, 5).value = "SEXO"
    sheet.cell(4, 6).value = "DNI"
    sheet.cell(4, 7).value = "CELULAR"
    sheet.cell(4, 8).value = "HOSPITAL"
    sheet.cell(4, 9).value = "DIRECCION"
    sheet.cell(4, 10).value = "HTA"
    sheet.cell(4, 11).value = "FECHA HTA"
    sheet.cell(4, 12).value = "DM"
    sheet.cell(4, 13).value = "FECHA DM"
    sheet.cell(4, 14).value = "HISTORIA"        
    sheet.cell(4, 15).value = "OBSERVACIÓN"

    for cell in sheet["4:4"]:
        cell.font = red_font


    for (i, paciente) in enumerate(hipertensos, 5): 
        sheet.cell(i, 1).value = i - 4
        sheet.cell(i, 2).value = paciente.nombre
        sheet.cell(i, 3).value = paciente.nacimiento
        sheet.cell(i, 4).value = codeHtml.ageInYears(tmp, paciente.nacimiento)
        sheet.cell(i, 5).value = "M" if paciente.sexo == 1 else "F"
        sheet.cell(i, 6).value = paciente.dni
        sheet.cell(i, 7).value = paciente.celular
        sheet.cell(i, 8).value = paciente.hospital.hospital
        sheet.cell(i, 9).value = paciente.direccion.title()
        sheet.cell(i, 10).value = "X" if paciente.hta else ""
        sheet.cell(i, 11).value = paciente.fechaHta if paciente.hta else ""
        sheet.cell(i, 12).value = "X" if paciente.dm else ""
        sheet.cell(i, 13).value = paciente.fechaDm if paciente.dm else ""
        sheet.cell(i, 14).value = paciente.historia        
        sheet.cell(i, 15).value = paciente.observacion

    sheet = wb.create_sheet(title = "DM")

    sheet.cell(2, 2).value = "PACIENTES CON DM"
    sheet.cell(2, 2).font = blue_font

    sheet.cell(4, 1).value = "#"
    sheet.cell(4, 2).value = "NOMBRE"
    sheet.cell(4, 3).value = "NACIMIENTO"
    sheet.cell(4, 4).value = "EDAD"
    sheet.cell(4, 5).value = "SEXO"
    sheet.cell(4, 6).value = "DNI"
    sheet.cell(4, 7).value = "CELULAR"
    sheet.cell(4, 8).value = "HOSPITAL"
    sheet.cell(4, 9).value = "DIRECCION"
    sheet.cell(4, 10).value = "HTA"
    sheet.cell(4, 11).value = "FECHA HTA"
    sheet.cell(4, 12).value = "DM"
    sheet.cell(4, 13).value = "FECHA DM"
    sheet.cell(4, 14).value = "HISTORIA"        
    sheet.cell(4, 15).value = "OBSERVACIÓN"

    for cell in sheet["4:4"]:
        cell.font = red_font
    
    for (i, paciente) in enumerate(deabeticos, 5): 
        sheet.cell(i, 1).value = i - 4
        sheet.cell(i, 2).value = paciente.nombre
        sheet.cell(i, 3).value = paciente.nacimiento
        sheet.cell(i, 4).value = codeHtml.ageInYears(tmp, paciente.nacimiento)
        sheet.cell(i, 5).value = "M" if paciente.sexo == 1 else "F"
        sheet.cell(i, 6).value = paciente.dni
        sheet.cell(i, 7).value = paciente.celular
        sheet.cell(i, 8).value = paciente.hospital.hospital
        sheet.cell(i, 9).value = paciente.direccion
        sheet.cell(i, 10).value = "X" if paciente.hta else ""
        sheet.cell(i, 11).value = paciente.fechaHta if paciente.hta else ""
        sheet.cell(i, 12).value = "X" if paciente.dm else ""
        sheet.cell(i, 13).value = paciente.fechaDm if paciente.dm else ""
        sheet.cell(i, 14).value = paciente.historia        
        sheet.cell(i, 15).value = paciente.observacion

    sheet = wb.create_sheet(title = "HTAxDM")

    sheet.cell(2, 2).value = "PACIENTES CON HTAxDM"
    sheet.cell(2, 2).font = blue_font

    sheet.cell(4, 1).value = "#"
    sheet.cell(4, 2).value = "NOMBRE"
    sheet.cell(4, 3).value = "NACIMIENTO"
    sheet.cell(4, 4).value = "EDAD"
    sheet.cell(4, 5).value = "SEXO"
    sheet.cell(4, 6).value = "DNI"
    sheet.cell(4, 7).value = "CELULAR"
    sheet.cell(4, 8).value = "HOSPITAL"
    sheet.cell(4, 9).value = "DIRECCION"
    sheet.cell(4, 10).value = "HTA"
    sheet.cell(4, 11).value = "FECHA HTA"
    sheet.cell(4, 12).value = "DM"
    sheet.cell(4, 13).value = "FECHA DM"
    sheet.cell(4, 14).value = "HISTORIA"        
    sheet.cell(4, 15).value = "OBSERVACIÓN"

    for cell in sheet["4:4"]:
        cell.font = red_font
    
    for (i, paciente) in enumerate(ambos, 5): 
        sheet.cell(i, 1).value = i - 4
        sheet.cell(i, 2).value = paciente.nombre
        sheet.cell(i, 3).value = paciente.nacimiento
        sheet.cell(i, 4).value = codeHtml.ageInYears(tmp, paciente.nacimiento)
        sheet.cell(i, 5).value = "M" if paciente.sexo == 1 else "F"
        sheet.cell(i, 6).value = paciente.dni
        sheet.cell(i, 7).value = paciente.celular
        sheet.cell(i, 8).value = paciente.hospital.hospital
        sheet.cell(i, 9).value = paciente.direccion
        sheet.cell(i, 10).value = "X" if paciente.hta else ""
        sheet.cell(i, 11).value = paciente.fechaHta if paciente.hta else ""
        sheet.cell(i, 12).value = "X" if paciente.dm else ""
        sheet.cell(i, 13).value = paciente.fechaDm if paciente.dm else ""
        sheet.cell(i, 14).value = paciente.historia        
        sheet.cell(i, 15).value = paciente.observacion
    
    # Save the workbook
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    
    response['Content-Disposition'] = "attachment; filename=pacientes.xlsx"

    wb.save(response)


    #wb.save(responde,title="pacientes.xlsx", overwrite = True)
    
    return response


def excel_02(request):
    tmp = date.today()

    _year = int(request.POST["nowYear"])

    red_font = Font(color = '00FF0000', bold = True)
    blue_font = Font(color = '00009900', bold = True, size=12)

    wb = Workbook()

    for _mes in range(1, 13):  

        hipertensos = Paciente.objects.filter(hta = True, fechaHta__month = _mes, fechaHta__year = _year ).order_by("nombre")
        deabeticos = Paciente.objects.filter(dm = True, fechaDm__month = _mes, fechaDm__year = _year ).order_by("nombre")

        counter = [0, 0]
    
        sheet = wb.create_sheet(title = str(Meses[_mes]))  
  
        
        sheet.cell(row=4, column=5).value = date.today()
        
        _last = 10

        sheet.cell(_last, 2).value = "PACIENTES CON HTA"
        sheet.cell(_last, 2).font = blue_font    

        _last += 2

        sheet.cell(_last, 1).value = "#"
        sheet.cell(_last, 2).value = "NOMBRE"
        sheet.cell(_last, 3).value = "NACIMIENTO"
        sheet.cell(_last, 4).value = "EDAD"
        sheet.cell(_last, 5).value = "SEXO"
        sheet.cell(_last, 6).value = "DNI"
        sheet.cell(_last, 7).value = "CELULAR"
        sheet.cell(_last, 8).value = "HOSPITAL"
        sheet.cell(_last, 9).value = "DIRECCION"
        sheet.cell(_last, 10).value = "HTA"
        sheet.cell(_last, 11).value = "FECHA HTA"
        sheet.cell(_last, 12).value = "DM"
        sheet.cell(_last, 13).value = "FECHA DM"
        sheet.cell(_last, 14).value = "HISTORIA"        
        sheet.cell(_last, 15).value = "OBSERVACIÓN"
        sheet.cell(_last, 16).value = "DETALLE"

        for cell in sheet[str(_last) + ":" + str(_last)]:
            cell.font = red_font

        _last += 1

        for (i, paciente) in enumerate(hipertensos): 
            sheet.cell(_last + i, 1).value = i + 1
            sheet.cell(_last + i, 2).value = paciente.nombre
            sheet.cell(_last + i, 3).value = paciente.nacimiento
            sheet.cell(_last + i, 4).value = codeHtml.ageInYears(tmp, paciente.nacimiento)
            sheet.cell(_last + i, 5).value = "M" if paciente.sexo == 1 else "F"
            sheet.cell(_last + i, 6).value = paciente.dni
            sheet.cell(_last + i, 7).value = paciente.celular
            sheet.cell(_last + i, 8).value = paciente.hospital.hospital
            sheet.cell(_last + i, 9).value = paciente.direccion.title()
            sheet.cell(_last + i, 10).value = "X" if paciente.hta else ""
            sheet.cell(_last + i, 11).value = paciente.fechaHta if paciente.hta else ""
            sheet.cell(_last + i, 12).value = "X" if paciente.dm else ""
            sheet.cell(_last + i, 13).value = paciente.fechaDm if paciente.dm else ""
            sheet.cell(_last + i, 14).value = paciente.historia        
            sheet.cell(_last + i, 15).value = paciente.observacion

            if paciente.dm == False: 
                sheet.cell(_last + i, 16).value = "NUEVO"
                counter[0] += 1 

            elif paciente.fechaDm.month == paciente.fechaHta.month and paciente.fechaDm.year == paciente.fechaHta.year:
                sheet.cell(_last + i, 16).value = "NUEVO"
                counter[0] += 1 

            elif paciente.fechaHta < paciente.fechaDm: 
                sheet.cell(_last + i, 16).value = "NUEVO"
                counter[0] += 1 

            else: 
                sheet.cell(_last + i, 16).value = "REINGRESANTE"









        _last += len(hipertensos) + 10

        sheet.cell(_last, 2).value = "PACIENTES CON DM"
        sheet.cell(_last, 2).font = blue_font    

        _last += 2

        sheet.cell(_last, 1).value = "#"
        sheet.cell(_last, 2).value = "NOMBRE"
        sheet.cell(_last, 3).value = "NACIMIENTO"
        sheet.cell(_last, 4).value = "EDAD"
        sheet.cell(_last, 5).value = "SEXO"
        sheet.cell(_last, 6).value = "DNI"
        sheet.cell(_last, 7).value = "CELULAR"
        sheet.cell(_last, 8).value = "HOSPITAL"
        sheet.cell(_last, 9).value = "DIRECCION"
        sheet.cell(_last, 10).value = "HTA"
        sheet.cell(_last, 11).value = "FECHA HTA"
        sheet.cell(_last, 12).value = "DM"
        sheet.cell(_last, 13).value = "FECHA DM"
        sheet.cell(_last, 14).value = "HISTORIA"        
        sheet.cell(_last, 15).value = "OBSERVACIÓN"
        sheet.cell(_last, 16).value = "DETALLE"

        for cell in sheet[str(_last) + ":" + str(_last)]:
            cell.font = red_font
        
        _last += 1

        for (i, paciente) in enumerate(deabeticos): 
            sheet.cell(_last + i, 1).value = i + 1
            sheet.cell(_last + i, 2).value = paciente.nombre
            sheet.cell(_last + i, 3).value = paciente.nacimiento
            sheet.cell(_last + i, 4).value = codeHtml.ageInYears(tmp, paciente.nacimiento)
            sheet.cell(_last + i, 5).value = "M" if paciente.sexo == 1 else "F"
            sheet.cell(_last + i, 6).value = paciente.dni
            sheet.cell(_last + i, 7).value = paciente.celular
            sheet.cell(_last + i, 8).value = paciente.hospital.hospital
            sheet.cell(_last + i, 9).value = paciente.direccion.title()
            sheet.cell(_last + i, 10).value = "X" if paciente.hta else ""
            sheet.cell(_last + i, 11).value = paciente.fechaHta if paciente.hta else ""
            sheet.cell(_last + i, 12).value = "X" if paciente.dm else ""
            sheet.cell(_last + i, 13).value = paciente.fechaDm if paciente.dm else ""
            sheet.cell(_last + i, 14).value = paciente.historia        
            sheet.cell(_last + i, 15).value = paciente.observacion

            if paciente.hta == False: 
                sheet.cell(_last + i, 16).value = "NUEVO"
                counter[1] += 1 

            elif paciente.fechaDm.month == paciente.fechaHta.month and paciente.fechaDm.year == paciente.fechaHta.year:
                sheet.cell(_last + i, 16).value = "NUEVO"
                counter[1] += 1 

            elif paciente.fechaHta > paciente.fechaDm: 
                sheet.cell(_last + i, 16).value = "NUEVO"
                counter[1] += 1 

            else: 
                sheet.cell(_last + i, 16).value = "REINGRESANTE"

        sheet.cell(row=5, column=2).font = red_font
        sheet.cell(row=5, column=2).value = "Pacientes con DM"    
        sheet.cell(row=5, column=3).value = counter[0]

        sheet.cell(row=6, column=2).font = red_font
        sheet.cell(row=6, column=2).value = "Pacientes con HTA"
        sheet.cell(row=6, column=3).value = counter[1]

        sheet.cell(row=7, column=2).font = red_font
        sheet.cell(row=7, column=2).value = "Total"
        sheet.cell(row=7, column=3).value = sum(counter)


    
    # Save the workbook
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    
    response['Content-Disposition'] = "attachment; filename=nuevos.xlsx"

    wb.save(response)


    #wb.save(responde,title="pacientes.xlsx", overwrite = True)
    
    return response


def excel_ekg(request):
    
    _year = int(request.POST["nowYear"])

    red_font = Font(color = '00FF0000', bold = True)
    blue_font = Font(color = '00009900', bold = True, size=12)

    wb = Workbook()

    for _mes in range(1, 13):    
         
        conLesion = Ekg.objects.filter(data__month = _mes, data__year = _year, esNormal = False).order_by("data")
        sinLesion = Ekg.objects.filter(data__month = _mes, data__year = _year, esNormal = True).order_by("data")

        sheet = wb.create_sheet(title = str(Meses[_mes]))        

        sheet.cell(row=3, column=3).value = "INFORME : " + Meses[_mes] + " - " + str(_year)
        sheet.cell(row=3, column=3).font = blue_font
        
        sheet.cell(row=5, column=3).font = red_font
        sheet.cell(row=5, column=3).value = "Cantidad total de EKGS NORMALES"                
        sheet.cell(row=5, column=4).value = len(sinLesion)

        sheet.cell(row=6, column=3).font = red_font
        sheet.cell(row=6, column=3).value = "Cantidad total de EKGS ALTERADOS"                        
        sheet.cell(row=6, column=4).value = len(conLesion)

        sheet.cell(row=7, column=3).font = red_font
        sheet.cell(row=7, column=3).value = "Cantidad total de EKGS realizados"                        
        sheet.cell(row=7, column=4).value = len(conLesion) + len(sinLesion)
        
        sheet.cell(row=10, column=3).value = "EKGs ALTERADOS"
        sheet.cell(row=10, column=3).font = blue_font

        _last = 12

        sheet.cell(row=_last, column=1).value = "#"
        sheet.cell(row=_last, column=2).value = "FECHA"
        sheet.cell(row=_last, column=3).value = "PACIENTE"
        sheet.cell(row=_last, column=4).value = "EDAD"
        sheet.cell(row=_last, column=5).value = "SEXO"
        sheet.cell(row=_last, column=6).value = "DX"
        sheet.cell(row=_last, column=7).value = "DNI"
        sheet.cell(row=_last, column=8).value = "HISTORIA"
        sheet.cell(row=_last, column=9).value = "ESTADO"        
        sheet.cell(row=_last, column=10).value = "OBSERVACION"

        for cell in sheet[str(_last) + ":" + str(_last)]:
            cell.font = red_font
        
        _last += 1 

        for (i, atencion) in enumerate(conLesion):
            sheet.cell(row=_last + i, column=1).value = i + 1
            sheet.cell(row=_last + i, column=2).value = atencion.data
            sheet.cell(row=_last + i, column=3).value = atencion.dni.nombre
            sheet.cell(row=_last + i, column=4).value = codeHtml.ageInYears(atencion.data, atencion.dni.nacimiento)
            sheet.cell(row=_last + i, column=5).value = "M" if atencion.dni.sexo else "F"
            tmp = " HTA" if atencion.dni.hta else ""
            tmp += " DM" if atencion.dni.dm else ""

            sheet.cell(row=_last + i, column=6).value = tmp
            sheet.cell(row=_last + i, column=7).value = atencion.dni.dni
            sheet.cell(row=_last + i, column=8).value = atencion.dni.historia
            sheet.cell(row=_last + i, column=9).value = "ALTERADO"
            sheet.cell(row=_last + i, column=10).value = atencion.observacion

        _last += len(conLesion) + 10

        sheet.cell(row=_last, column=3).value = "EKGs Normales"
        sheet.cell(row=_last, column=3).font = blue_font

        _last += 2

        sheet.cell(row=_last, column=1).value = "#"
        sheet.cell(row=_last, column=2).value = "FECHA"
        sheet.cell(row=_last, column=3).value = "PACIENTE"
        sheet.cell(row=_last, column=4).value = "EDAD"
        sheet.cell(row=_last, column=5).value = "SEXO"
        sheet.cell(row=_last, column=6).value = "DX"
        sheet.cell(row=_last, column=7).value = "DNI"
        sheet.cell(row=_last, column=8).value = "HISTORIA"
        sheet.cell(row=_last, column=9).value = "ESTADO"        
        sheet.cell(row=_last, column=10).value = "OBSERVACION"

        for cell in sheet[str(_last) + ":" + str(_last)]:
            cell.font = red_font
        
        _last += 1 

        for (i, atencion) in enumerate(sinLesion):
            sheet.cell(row=_last + i, column=1).value = i + 1
            sheet.cell(row=_last + i, column=2).value = atencion.data
            sheet.cell(row=_last + i, column=3).value = atencion.dni.nombre
            sheet.cell(row=_last + i, column=4).value = codeHtml.ageInYears(atencion.data, atencion.dni.nacimiento)
            sheet.cell(row=_last + i, column=5).value = "M" if atencion.dni.sexo else "F"
            tmp = " HTA" if atencion.dni.hta else ""
            tmp += " DM" if atencion.dni.dm else ""

            sheet.cell(row=_last + i, column=6).value = tmp
            sheet.cell(row=_last + i, column=7).value = atencion.dni.dni
            sheet.cell(row=_last + i, column=8).value = atencion.dni.historia
            sheet.cell(row=_last + i, column=9).value = "NORMAL"
            sheet.cell(row=_last + i, column=10).value = atencion.observacion
       

    # Save the workbook
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    
    response['Content-Disposition'] = "attachment; filename=Ekgs.xlsx"

    wb.save(response)

    return response

def excel_pie(request):
    _year = int(request.POST["nowYear"])

    red_font = Font(color = '00FF0000', bold = True)
    blue_font = Font(color = '00009900', bold = True, size=12)

    wb = Workbook()

    for _mes in range(1, 13):    
         
        conLesion = Pie.objects.filter(data__month = _mes, data__year = _year, conLesion = True).order_by("data")
        sinLesion = Pie.objects.filter(data__month = _mes, data__year = _year, conLesion = False).order_by("data")
        
        sheet = wb.create_sheet(title = str(Meses[_mes]))  

        sheet.cell(row=3, column=3).value = "INFORME : " + Meses[_mes] + " - " + str(_year)
        sheet.cell(row=3, column=3).font = blue_font
        
        sheet.cell(row=5, column=3).font = red_font
        sheet.cell(row=5, column=3).value = "Cantidad total de CHEQUEOS SIN LESSIÓN"                
        sheet.cell(row=5, column=4).value = len(sinLesion)

        sheet.cell(row=6, column=3).font = red_font
        sheet.cell(row=6, column=3).value = "Cantidad total de CHEQUEOS CON LESSIÓN"                        
        sheet.cell(row=6, column=4).value = len(conLesion)

        sheet.cell(row=7, column=3).font = red_font
        sheet.cell(row=7, column=3).value = "Cantidad total de CHEQUEOS PE PIES"                        
        sheet.cell(row=7, column=4).value = len(conLesion) + len(sinLesion)
        
        sheet.cell(row=10, column=3).value = "PIES CON LESIÓN"
        sheet.cell(row=10, column=3).font = blue_font

        _last = 12

        sheet.cell(row=_last, column=1).value = "#"
        sheet.cell(row=_last, column=2).value = "FECHA"
        sheet.cell(row=_last, column=3).value = "PACIENTE"
        sheet.cell(row=_last, column=4).value = "EDAD"
        sheet.cell(row=_last, column=5).value = "SEXO"
        sheet.cell(row=_last, column=6).value = "DX"
        sheet.cell(row=_last, column=7).value = "DNI"
        sheet.cell(row=_last, column=8).value = "HISTORIA"
        sheet.cell(row=_last, column=9).value = "ESTADO"        
        sheet.cell(row=_last, column=10).value = "OBSERVACION"

        for cell in sheet[str(_last) + ":" + str(_last)]:
            cell.font = red_font
        
        _last += 1 

        for (i, atencion) in enumerate(conLesion):
            sheet.cell(row=_last + i, column=1).value = i + 1
            sheet.cell(row=_last + i, column=2).value = atencion.data
            sheet.cell(row=_last + i, column=3).value = atencion.dni.nombre
            sheet.cell(row=_last + i, column=4).value = codeHtml.ageInYears(atencion.data, atencion.dni.nacimiento)
            sheet.cell(row=_last + i, column=5).value = "M" if atencion.dni.sexo else "F"
            tmp = " HTA" if atencion.dni.hta else ""
            tmp += " DM" if atencion.dni.dm else ""

            sheet.cell(row=_last + i, column=6).value = tmp
            sheet.cell(row=_last + i, column=7).value = atencion.dni.dni
            sheet.cell(row=_last + i, column=8).value = atencion.dni.historia
            sheet.cell(row=_last + i, column=9).value = "Con Lesion"
            sheet.cell(row=_last + i, column=10).value = atencion.observacion

        _last += len(conLesion) + 10

        sheet.cell(row=_last, column=3).value = "SIN LESIÓN"
        sheet.cell(row=_last, column=3).font = blue_font

        _last += 2

        sheet.cell(row=_last, column=1).value = "#"
        sheet.cell(row=_last, column=2).value = "FECHA"
        sheet.cell(row=_last, column=3).value = "PACIENTE"
        sheet.cell(row=_last, column=4).value = "EDAD"
        sheet.cell(row=_last, column=5).value = "SEXO"
        sheet.cell(row=_last, column=6).value = "DX"
        sheet.cell(row=_last, column=7).value = "DNI"
        sheet.cell(row=_last, column=8).value = "HISTORIA"
        sheet.cell(row=_last, column=9).value = "ESTADO"        
        sheet.cell(row=_last, column=10).value = "OBSERVACION"

        for cell in sheet[str(_last) + ":" + str(_last)]:
            cell.font = red_font
        
        _last += 1 

        for (i, atencion) in enumerate(sinLesion):
            sheet.cell(row=_last + i, column=1).value = i + 1
            sheet.cell(row=_last + i, column=2).value = atencion.data
            sheet.cell(row=_last + i, column=3).value = atencion.dni.nombre
            sheet.cell(row=_last + i, column=4).value = codeHtml.ageInYears(atencion.data, atencion.dni.nacimiento)
            sheet.cell(row=_last + i, column=5).value = "M" if atencion.dni.sexo else "F"
            tmp = " HTA" if atencion.dni.hta else ""
            tmp += " DM" if atencion.dni.dm else ""

            sheet.cell(row=_last + i, column=6).value = tmp
            sheet.cell(row=_last + i, column=7).value = atencion.dni.dni
            sheet.cell(row=_last + i, column=8).value = atencion.dni.historia
            sheet.cell(row=_last + i, column=9).value = "Sin Lesion"
            sheet.cell(row=_last + i, column=10).value = atencion.observacion
       

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    
    response['Content-Disposition'] = "attachment; filename=Pies.xlsx"

    wb.save(response)

    return response



def excel_dot(request):
    _year = int(request.POST["nowYear"])

    red_font = Font(color = '00FF0000', bold = True)
    blue_font = Font(color = '00009900', bold = True, size=12)

    wb = Workbook()

    for _mes in range(1, 13):    

        counter = [0, 0]
         
        dot1 = Atencion.objects.filter(fecha__month = _mes, fecha__year = _year, dotacion = "1").order_by("fecha")
        dot2 = Atencion.objects.filter(fecha__month = _mes, fecha__year = _year, dotacion = "2").order_by("fecha")
        dot3 = Atencion.objects.filter(fecha__month = _mes, fecha__year = _year, dotacion = "3").order_by("fecha")
        dot4 = Atencion.objects.filter(fecha__month = _mes, fecha__year = _year, dotacion = "4").order_by("fecha")
    
        sheet = wb.create_sheet(title = str(Meses[_mes]))    

        sheet.cell(row=3, column=3).value = "INFORME : " + str(Meses[_mes]) + " - " + str(_year)
        sheet.cell(row=3, column=3).font = blue_font

        sheet.cell(row=5, column=3).font = red_font
        sheet.cell(row=5, column=3).value = "Cantidad total de atenciones DOTACIÓN I"        
        sheet.cell(row=5, column=4).value = len(dot1)

        sheet.cell(row=6, column=3).font = red_font
        sheet.cell(row=6, column=3).value = "Cantidad total de atenciones DOTACIÓN II"                
        sheet.cell(row=6, column=4).value = len(dot2)

        sheet.cell(row=7, column=3).font = red_font
        sheet.cell(row=7, column=3).value = "Cantidad total de atenciones DOTACIÓN III"                
        sheet.cell(row=7, column=4).value = len(dot3)

        sheet.cell(row=8, column=3).font = red_font
        sheet.cell(row=8, column=3).value = "Cantidad total de atenciones OTROS"                
        sheet.cell(row=8, column=4).value = len(dot4)

        sheet.cell(row=4, column=5).font = red_font
        sheet.cell(row=4, column=5).value = "Nuevos"                
        sheet.cell(row=4, column=6).font = red_font
        sheet.cell(row=4, column=6).value = "Reingresantes"                
        
        sheet.cell(row=10, column=3).value = "DOTACION I"
        sheet.cell(row=10, column=3).font = blue_font

        _last = 12

        sheet.cell(row=_last, column=1).value = "#"
        sheet.cell(row=_last, column=2).value = "FECHA"
        sheet.cell(row=_last, column=3).value = "PACIENTE"
        sheet.cell(row=_last, column=4).value = "EDAD"
        sheet.cell(row=_last, column=5).value = "SEXO"
        sheet.cell(row=_last, column=6).value = "DX"
        sheet.cell(row=_last, column=7).value = "DNI"
        sheet.cell(row=_last, column=8).value = "HISTORIA"
        sheet.cell(row=_last, column=9).value = "PESO"
        sheet.cell(row=_last, column=10).value = "TALLA"
        sheet.cell(row=_last, column=11).value = "P. ABDOMINAL"
        sheet.cell(row=_last, column=12).value = "PRESIÓN A."
        sheet.cell(row=_last, column=13).value = "GLUCOTEST"
        sheet.cell(row=_last, column=14).value = "DETALLE"

        for cell in sheet[str(_last) + ":" + str(_last)]:
            cell.font = red_font
        
        _last += 1 

        for (i, atencion) in enumerate(dot1):
            sheet.cell(row=_last + i, column=1).value = i + 1
            sheet.cell(row=_last + i, column=2).value = atencion.fecha
            sheet.cell(row=_last + i, column=3).value = atencion.dni.nombre
            sheet.cell(row=_last + i, column=4).value = codeHtml.ageInYears(atencion.fecha, atencion.dni.nacimiento)
            sheet.cell(row=_last + i, column=5).value = "M" if atencion.dni.sexo else "F"
            tmp = " HTA" if atencion.dni.hta else ""
            tmp += " DM" if atencion.dni.dm else ""

            sheet.cell(row=_last + i, column=6).value = tmp
            sheet.cell(row=_last + i, column=7).value = atencion.dni.dni
            sheet.cell(row=_last + i, column=8).value = atencion.dni.historia
            sheet.cell(row=_last + i, column=9).value = atencion.peso
            sheet.cell(row=_last + i, column=10).value = atencion.talla
            sheet.cell(row=_last + i, column=11).value = atencion.perimetro
            sheet.cell(row=_last + i, column=12).value = atencion.presion
            sheet.cell(row=_last + i, column=13).value = str(atencion.hemotest) + " MMHG"
            
            tmp = None 
            if atencion.dni.dm: 
                tmp = atencion.dni.fechaDm
            
            if atencion.dni.hta: 
                if tmp == None:
                    tmp = atencion.dni.fechaHta

                else:
                    tmp = min(tmp, atencion.dni.fechaHta) 

            if tmp.month == atencion.fecha.month and tmp.year == atencion.fecha.year: 
                sheet.cell(row=_last + i, column=14).value = "NUEVO"

                counter[0] += 1
            
            else: 
                sheet.cell(row=_last + i, column=14).value = "REINGRESANTE"

                counter[1] += 1


        sheet.cell(row=5, column=5).value = counter[0]
        sheet.cell(row=5, column=6).value = counter[1]

        _last += len(dot1) + 10

        sheet.cell(row=_last, column=3).value = "DOTACION II"
        sheet.cell(row=_last, column=3).font = blue_font

        _last += 2

        sheet.cell(row=_last, column=1).value = "#"
        sheet.cell(row=_last, column=2).value = "FECHA"
        sheet.cell(row=_last, column=3).value = "PACIENTE"
        sheet.cell(row=_last, column=4).value = "EDAD"
        sheet.cell(row=_last, column=5).value = "SEXO"
        sheet.cell(row=_last, column=6).value = "DX"
        sheet.cell(row=_last, column=7).value = "DNI"
        sheet.cell(row=_last, column=8).value = "HISTORIA"
        sheet.cell(row=_last, column=9).value = "PESO"
        sheet.cell(row=_last, column=10).value = "TALLA"
        sheet.cell(row=_last, column=11).value = "P. ABDOMINAL"
        sheet.cell(row=_last, column=12).value = "PRESIÓN A."
        sheet.cell(row=_last, column=13).value = "GLUCOTEST"

        for cell in sheet[str(_last) + ":" + str(_last)]:
            cell.font = red_font

        _last += 1 

        for (i, atencion) in enumerate(dot2):
            sheet.cell(row=_last + i, column=1).value = i + 1
            sheet.cell(row=_last + i, column=2).value = atencion.fecha
            sheet.cell(row=_last + i, column=3).value = atencion.dni.nombre
            sheet.cell(row=_last + i, column=4).value = codeHtml.ageInYears(atencion.fecha, atencion.dni.nacimiento)
            sheet.cell(row=_last + i, column=5).value = "M" if atencion.dni.sexo else "F"
            tmp = " HTA" if atencion.dni.hta else ""
            tmp += " DM" if atencion.dni.dm else ""

            sheet.cell(row=_last + i, column=6).value = tmp
            sheet.cell(row=_last + i, column=7).value = atencion.dni.dni
            sheet.cell(row=_last + i, column=8).value = atencion.dni.historia
            sheet.cell(row=_last + i, column=6).value = tmp
            sheet.cell(row=_last + i, column=7).value = atencion.dni.dni
            sheet.cell(row=_last + i, column=8).value = atencion.dni.historia
            sheet.cell(row=_last + i, column=9).value = atencion.peso
            sheet.cell(row=_last + i, column=10).value = atencion.talla
            sheet.cell(row=_last + i, column=11).value = atencion.perimetro
            sheet.cell(row=_last + i, column=12).value = atencion.presion
            sheet.cell(row=_last + i, column=13).value = str(atencion.hemotest) + " MMHG"


        _last += len(dot2) + 10

        sheet.cell(row=_last, column=3).value = "DOTACION III"
        sheet.cell(row=_last, column=3).font = blue_font


        _last += 2

        sheet.cell(row=_last, column=1).value = "#"
        sheet.cell(row=_last, column=2).value = "FECHA"
        sheet.cell(row=_last, column=3).value = "PACIENTE"
        sheet.cell(row=_last, column=4).value = "EDAD"
        sheet.cell(row=_last, column=5).value = "SEXO"
        sheet.cell(row=_last, column=6).value = "DX"
        sheet.cell(row=_last, column=7).value = "DNI"
        sheet.cell(row=_last, column=8).value = "HISTORIA"
        sheet.cell(row=_last, column=9).value = "PESO"
        sheet.cell(row=_last, column=10).value = "TALLA"
        sheet.cell(row=_last, column=11).value = "P. ABDOMINAL"
        sheet.cell(row=_last, column=12).value = "PRESIÓN A."
        sheet.cell(row=_last, column=13).value = "GLUCOTEST"

        for cell in sheet[str(_last) + ":" + str(_last)]:
            cell.font = red_font
        
        _last += 1 

        for (i, atencion) in enumerate(dot3):
            sheet.cell(row=_last + i, column=1).value = i + 1
            sheet.cell(row=_last + i, column=2).value = atencion.fecha
            sheet.cell(row=_last + i, column=3).value = atencion.dni.nombre
            sheet.cell(row=_last + i, column=4).value = codeHtml.ageInYears(atencion.fecha, atencion.dni.nacimiento)
            sheet.cell(row=_last + i, column=5).value = "M" if atencion.dni.sexo else "F"
            tmp = " HTA" if atencion.dni.hta else ""
            tmp += " DM" if atencion.dni.dm else ""

            sheet.cell(row=_last + i, column=6).value = tmp
            sheet.cell(row=_last + i, column=7).value = atencion.dni.dni
            sheet.cell(row=_last + i, column=8).value = atencion.dni.historia
            sheet.cell(row=_last + i, column=9).value = atencion.peso
            sheet.cell(row=_last + i, column=10).value = atencion.talla
            sheet.cell(row=_last + i, column=11).value = atencion.perimetro
            sheet.cell(row=_last + i, column=12).value = atencion.presion
            sheet.cell(row=_last + i, column=13).value = str(atencion.hemotest) + " MMHG"

        _last += len(dot3) + 10

        sheet.cell(row=_last, column=3).value = "OTROS"
        sheet.cell(row=_last, column=3).font = blue_font


        _last += 2

        sheet.cell(row=_last, column=1).value = "#"
        sheet.cell(row=_last, column=2).value = "FECHA"
        sheet.cell(row=_last, column=3).value = "PACIENTE"
        sheet.cell(row=_last, column=4).value = "EDAD"
        sheet.cell(row=_last, column=5).value = "SEXO"
        sheet.cell(row=_last, column=6).value = "DX"
        sheet.cell(row=_last, column=7).value = "DNI"
        sheet.cell(row=_last, column=8).value = "HISTORIA"
        sheet.cell(row=_last, column=9).value = "PESO"
        sheet.cell(row=_last, column=10).value = "TALLA"
        sheet.cell(row=_last, column=11).value = "P. ABDOMINAL"
        sheet.cell(row=_last, column=12).value = "PRESIÓN A."
        sheet.cell(row=_last, column=13).value = "GLUCOTEST"

        for cell in sheet[str(_last) + ":" + str(_last)]:
            cell.font = red_font
        
        _last += 1 

        for (i, atencion) in enumerate(dot4):
            sheet.cell(row=_last + i, column=1).value = i + 1
            sheet.cell(row=_last + i, column=2).value = atencion.fecha
            sheet.cell(row=_last + i, column=3).value = atencion.dni.nombre
            sheet.cell(row=_last + i, column=4).value = codeHtml.ageInYears(atencion.fecha, atencion.dni.nacimiento)
            sheet.cell(row=_last + i, column=5).value = "M" if atencion.dni.sexo else "F"
            tmp = " HTA" if atencion.dni.hta else ""
            tmp += " DM" if atencion.dni.dm else ""

            sheet.cell(row=_last + i, column=6).value = tmp
            sheet.cell(row=_last + i, column=7).value = atencion.dni.dni
            sheet.cell(row=_last + i, column=8).value = atencion.dni.historia
            sheet.cell(row=_last + i, column=9).value = atencion.peso
            sheet.cell(row=_last + i, column=10).value = atencion.talla
            sheet.cell(row=_last + i, column=11).value = atencion.perimetro
            sheet.cell(row=_last + i, column=12).value = atencion.presion
            sheet.cell(row=_last + i, column=13).value = str(atencion.hemotest) + " MMHG"

        _last += len(dot2) + 1


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    
    response['Content-Disposition'] = "attachment; filename=dotaciones.xlsx"

    wb.save(response)

    return response


def excel_imc(request):
    _year = int(request.POST["nowYear"])

    red_font = Font(color = '00FF0000', bold = True)
    blue_font = Font(color = '00009900', bold = True, size=12)

    wb = Workbook()

    for _mes in range(1, 13):    
         
        dot1 = Atencion.objects.filter(fecha__month = _mes, fecha__year = _year, imcDescripcion = "Normal").order_by("fecha")
        dot2 = Atencion.objects.filter(fecha__month = _mes, fecha__year = _year, imcDescripcion = "Sobrepeso").order_by("fecha")
        dot3 = Atencion.objects.filter(fecha__month = _mes, fecha__year = _year, imcDescripcion = "Obeso").order_by("fecha")
       
        sheet = wb.create_sheet(title = str(Meses[_mes]))    

        sheet.cell(row=3, column=3).value = "INFORME : " + str(Meses[_mes]) + " - " + str(_year)
        sheet.cell(row=3, column=3).font = blue_font

        sheet.cell(row=5, column=3).font = red_font
        sheet.cell(row=5, column=3).value = "Cantidad de pacientes con peso NORMAL"        
        sheet.cell(row=5, column=4).value = len(dot1)

        sheet.cell(row=6, column=3).font = red_font
        sheet.cell(row=6, column=3).value = "Cantidad de pacientes con peso SOBREPESO"                
        sheet.cell(row=6, column=4).value = len(dot2)

        sheet.cell(row=7, column=3).font = red_font
        sheet.cell(row=7, column=3).value = "Cantidad de pacientes con peso OBESO"                
        sheet.cell(row=7, column=4).value = len(dot3)

        sheet.cell(row=8, column=3).font = red_font
        sheet.cell(row=8, column=3).value = "Cantidad total de pacientes con análisis de peso: "                
        sheet.cell(row=8, column=4).value = len(dot1) + len(dot2) + len(dot3)

        sheet.cell(row=10, column=3).value = "NORMAL"
        sheet.cell(row=10, column=3).font = blue_font

        _last = 12

        sheet.cell(row=_last, column=1).value = "#"
        sheet.cell(row=_last, column=2).value = "FECHA"
        sheet.cell(row=_last, column=3).value = "PACIENTE"
        sheet.cell(row=_last, column=4).value = "EDAD"
        sheet.cell(row=_last, column=5).value = "SEXO"
        sheet.cell(row=_last, column=6).value = "DX"
        sheet.cell(row=_last, column=7).value = "DNI"
        sheet.cell(row=_last, column=8).value = "HISTORIA"
        sheet.cell(row=_last, column=9).value = "PESO"
        sheet.cell(row=_last, column=10).value = "TALLA"
        sheet.cell(row=_last, column=11).value = "PESO"
        sheet.cell(row=_last, column=12).value = "P. ABDOMINAL"
        sheet.cell(row=_last, column=13).value = "PRESIÓN A."
        sheet.cell(row=_last, column=14).value = "GLUCOTEST"

        for cell in sheet[str(_last) + ":" + str(_last)]:
            cell.font = red_font
        
        _last += 1 

        for (i, atencion) in enumerate(dot1):
            sheet.cell(row=_last + i, column=1).value = i + 1
            sheet.cell(row=_last + i, column=2).value = atencion.fecha
            sheet.cell(row=_last + i, column=3).value = atencion.dni.nombre
            sheet.cell(row=_last + i, column=4).value = codeHtml.ageInYears(atencion.fecha, atencion.dni.nacimiento)
            sheet.cell(row=_last + i, column=5).value = "M" if atencion.dni.sexo else "F"
            tmp = " HTA" if atencion.dni.hta else ""
            tmp += " DM" if atencion.dni.dm else ""

            sheet.cell(row=_last + i, column=6).value = tmp
            sheet.cell(row=_last + i, column=7).value = atencion.dni.dni
            sheet.cell(row=_last + i, column=8).value = atencion.dni.historia
            sheet.cell(row=_last + i, column=9).value = atencion.peso
            sheet.cell(row=_last + i, column=10).value = atencion.talla
            sheet.cell(row=_last + i, column=11).value = atencion.imcDescripcion
            sheet.cell(row=_last + i, column=12).value = atencion.perimetro
            sheet.cell(row=_last + i, column=13).value = atencion.presion
            sheet.cell(row=_last + i, column=14).value = str(atencion.hemotest) + " MMHG"

        _last += len(dot1) + 10

        sheet.cell(row=_last, column=3).value = "SOBREPESO"
        sheet.cell(row=_last, column=3).font = blue_font

        _last += 2

        
        sheet.cell(row=_last, column=1).value = "#"
        sheet.cell(row=_last, column=2).value = "FECHA"
        sheet.cell(row=_last, column=3).value = "PACIENTE"
        sheet.cell(row=_last, column=4).value = "EDAD"
        sheet.cell(row=_last, column=5).value = "SEXO"
        sheet.cell(row=_last, column=6).value = "DX"
        sheet.cell(row=_last, column=7).value = "DNI"
        sheet.cell(row=_last, column=8).value = "HISTORIA"
        sheet.cell(row=_last, column=9).value = "PESO"
        sheet.cell(row=_last, column=10).value = "TALLA"
        sheet.cell(row=_last, column=11).value = "PESO"
        sheet.cell(row=_last, column=12).value = "P. ABDOMINAL"
        sheet.cell(row=_last, column=13).value = "PRESIÓN A."
        sheet.cell(row=_last, column=14).value = "GLUCOTEST"

        for cell in sheet[str(_last) + ":" + str(_last)]:
            cell.font = red_font

        _last += 1 

        for (i, atencion) in enumerate(dot2):
            sheet.cell(row=_last + i, column=1).value = i + 1
            sheet.cell(row=_last + i, column=2).value = atencion.fecha
            sheet.cell(row=_last + i, column=3).value = atencion.dni.nombre
            sheet.cell(row=_last + i, column=4).value = codeHtml.ageInYears(atencion.fecha, atencion.dni.nacimiento)
            sheet.cell(row=_last + i, column=5).value = "M" if atencion.dni.sexo else "F"
            tmp = " HTA" if atencion.dni.hta else ""
            tmp += " DM" if atencion.dni.dm else ""

            sheet.cell(row=_last + i, column=6).value = tmp
            sheet.cell(row=_last + i, column=7).value = atencion.dni.dni
            sheet.cell(row=_last + i, column=8).value = atencion.dni.historia
            sheet.cell(row=_last + i, column=9).value = atencion.peso
            sheet.cell(row=_last + i, column=10).value = atencion.talla
            sheet.cell(row=_last + i, column=11).value = atencion.imcDescripcion
            sheet.cell(row=_last + i, column=12).value = atencion.perimetro
            sheet.cell(row=_last + i, column=13).value = atencion.presion
            sheet.cell(row=_last + i, column=14).value = str(atencion.hemotest) + " MMHG"

        _last += len(dot2) + 10

        sheet.cell(row=_last, column=3).value = "OBESO"
        sheet.cell(row=_last, column=3).font = blue_font

        _last += 2
        
        sheet.cell(row=_last, column=1).value = "#"
        sheet.cell(row=_last, column=2).value = "FECHA"
        sheet.cell(row=_last, column=3).value = "PACIENTE"
        sheet.cell(row=_last, column=4).value = "EDAD"
        sheet.cell(row=_last, column=5).value = "SEXO"
        sheet.cell(row=_last, column=6).value = "DX"
        sheet.cell(row=_last, column=7).value = "DNI"
        sheet.cell(row=_last, column=8).value = "HISTORIA"
        sheet.cell(row=_last, column=9).value = "PESO"
        sheet.cell(row=_last, column=10).value = "TALLA"
        sheet.cell(row=_last, column=11).value = "PESO"
        sheet.cell(row=_last, column=12).value = "P. ABDOMINAL"
        sheet.cell(row=_last, column=13).value = "PRESIÓN A."
        sheet.cell(row=_last, column=14).value = "GLUCOTEST"

        for cell in sheet[str(_last) + ":" + str(_last)]:
            cell.font = red_font
        
        _last += 1 

        for (i, atencion) in enumerate(dot3):
            sheet.cell(row=_last + i, column=1).value = i + 1
            sheet.cell(row=_last + i, column=2).value = atencion.fecha
            sheet.cell(row=_last + i, column=3).value = atencion.dni.nombre
            sheet.cell(row=_last + i, column=4).value = codeHtml.ageInYears(atencion.fecha, atencion.dni.nacimiento)
            sheet.cell(row=_last + i, column=5).value = "M" if atencion.dni.sexo else "F"
            tmp = " HTA" if atencion.dni.hta else ""
            tmp += " DM" if atencion.dni.dm else ""

            sheet.cell(row=_last + i, column=6).value = tmp
            sheet.cell(row=_last + i, column=7).value = atencion.dni.dni
            sheet.cell(row=_last + i, column=8).value = atencion.dni.historia
            sheet.cell(row=_last + i, column=9).value = atencion.peso
            sheet.cell(row=_last + i, column=10).value = atencion.talla
            sheet.cell(row=_last + i, column=11).value = atencion.imcDescripcion
            sheet.cell(row=_last + i, column=12).value = atencion.perimetro
            sheet.cell(row=_last + i, column=13).value = atencion.presion
            sheet.cell(row=_last + i, column=14).value = str(atencion.hemotest) + " MMHG"

        _last += len(dot3) + 10

       
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    
    response['Content-Disposition'] = "attachment; filename=obesidad.xlsx"

    wb.save(response)

    return response



def excel_ref(request):
    _year = int(request.POST["nowYear"])

    red_font = Font(color = '00FF0000', bold = True)
    blue_font = Font(color = '0000FF00', bold = True, size=12)
    yellow_font = Font(color = 'FF000000', italic = True)

    wb = Workbook()

    especialidad = Especialidad.objects.all()

    for _mes in range(1, 13):    
        sheet = wb.create_sheet(title = str(Meses[_mes])) 

        sheet.cell(row=3, column=3).value = "INFORME : " + str(Meses[_mes]) + " - " + str(_year)
        sheet.cell(row=3, column=3).font = blue_font 

        dot = {}

        _last = 10

        for (i, _esp) in enumerate(especialidad): 
            dot[_esp.id] = Referencia.objects.filter(data__month = _mes, data__year = _year, especialidad = _esp).order_by("data")
        
            sheet.cell(row=i + 5, column=3).font = red_font
            sheet.cell(row=i + 5, column=3).value = "Cantidad de referencias a "+ _esp.especialidad        
            sheet.cell(row=i + 5, column=4).value = len(dot[_esp.id])

        _last = 5 + len(dot) + 3
        for (i, _esp) in enumerate(especialidad):            
            sheet.cell(row=_last, column=3).value = _esp.especialidad.upper()
            sheet.cell(row=_last, column=3).font = blue_font

            _last += 2

            sheet.cell(row=_last, column=1).value = "#"
            sheet.cell(row=_last, column=2).value = "FECHA"
            sheet.cell(row=_last, column=3).value = "PACIENTE"
            sheet.cell(row=_last, column=4).value = "EDAD"
            sheet.cell(row=_last, column=5).value = "SEXO"
            sheet.cell(row=_last, column=6).value = "DX"
            sheet.cell(row=_last, column=7).value = "DNI"
            sheet.cell(row=_last, column=8).value = "HISTORIA"
            sheet.cell(row=_last, column=9).value = "REFERENCIA"
            sheet.cell(row=_last, column=10).value = "ATENDIDA"
            sheet.cell(row=_last, column=11).value = "OBSERVACIÓN"
            
            for cell in sheet[str(_last) + ":" + str(_last)]:
                cell.font = red_font
        
            _last += 1 

            for (i, atencion) in enumerate(dot[_esp.id]):
                sheet.cell(row=_last + i, column=1).value = i + 1
                sheet.cell(row=_last + i, column=2).value = atencion.data
                sheet.cell(row=_last + i, column=3).value = atencion.dni.nombre
                sheet.cell(row=_last + i, column=4).value = codeHtml.ageInYears(atencion.data, atencion.dni.nacimiento)
                sheet.cell(row=_last + i, column=5).value = "M" if atencion.dni.sexo else "F"
                tmp = " HTA" if atencion.dni.hta else ""
                tmp += " DM" if atencion.dni.dm else ""

                sheet.cell(row=_last + i, column=6).value = tmp
                sheet.cell(row=_last + i, column=7).value = atencion.dni.dni
                sheet.cell(row=_last + i, column=8).value = atencion.dni.historia
                sheet.cell(row=_last + i, column=9).value = atencion.especialidad.especialidad
                sheet.cell(row=_last + i, column=10).value = "Atendido" if atencion.atendido else ""
                sheet.cell(row=_last + i, column=11).value = atencion.observacion

       
            _last += len(dot[_esp.id]) + 10

       
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    
    response['Content-Disposition'] = "attachment; filename=referencias.xlsx"

    wb.save(response)

    return response

def excel_dead(request):
    _year = int(request.POST["nowYear"])

    tmp = date.today()

    muertos = Paciente.objects.filter(muerto = True, fechaMuerto__year = _year).order_by("fechaMuerto")

    red_font = Font(color = '00FF0000', bold = True)
    blue_font = Font(color = '00009900', bold = True, size=12)

    wb = Workbook()

    sheet = wb.active
    sheet.title = "muertos"

    sheet.cell(row=2, column=2).font = red_font
    sheet.cell(row=2, column=2).value = "Pacientes muertos en el " + str(_year)
    sheet.cell(row=2, column=3).value = len(muertos)
    
    sheet.cell(4, 1).value = "#"
    sheet.cell(4, 2).value = "NOMBRE"
    sheet.cell(4, 3).value = "NACIMIENTO"
    sheet.cell(4, 4).value = "EDAD"
    sheet.cell(4, 5).value = "SEXO"
    sheet.cell(4, 6).value = "DNI"
    sheet.cell(4, 7).value = "CELULAR"
    sheet.cell(4, 8).value = "HOSPITAL"
    sheet.cell(4, 9).value = "DIRECCION"
    sheet.cell(4, 10).value = "DIAGNÓSTICOS"
    sheet.cell(4, 11).value = "HISTORIA"        
    sheet.cell(4, 12).value = "OBSERVACIÓN"
    sheet.cell(4, 13).value = "DECESO"
    sheet.cell(4, 14).value = "FECHA-DECESO"
    sheet.cell(4, 15).value = "OBSERVACIÓN-DECESO"
    

    for cell in sheet["4:4"]:
        cell.font = red_font


    for (i, paciente) in enumerate(muertos, 5): 
        sheet.cell(i, 1).value = i - 4
        sheet.cell(i, 2).value = paciente.nombre
        sheet.cell(i, 3).value = paciente.nacimiento
        sheet.cell(i, 4).value = codeHtml.ageInYears(paciente.fechaMuerto, paciente.nacimiento)
        sheet.cell(i, 5).value = "M" if paciente.sexo == 1 else "F"
        sheet.cell(i, 6).value = paciente.dni
        sheet.cell(i, 7).value = paciente.celular
        sheet.cell(i, 8).value = paciente.hospital.hospital
        sheet.cell(i, 9).value = paciente.direccion.title()
        
        tmp = " HTA" if paciente.hta else ""
        tmp += " DM" if paciente.hta else ""        

        sheet.cell(i, 10).value = tmp
        
        sheet.cell(i, 11).value = paciente.historia
        sheet.cell(i, 12).value = paciente.observacion
        sheet.cell(i, 13).value = "MUERTO" if paciente.muerto else ""
        sheet.cell(i, 14).value = paciente.fechaMuerto
        sheet.cell(i, 15).value = paciente.observacionMuerto

    # Save the workbook
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    
    response['Content-Disposition'] = "attachment; filename=muertos.xlsx"

    wb.save(response)


    #wb.save(responde,title="pacientes.xlsx", overwrite = True)
    
    return response


























def super_admin(request):
    if "texto" in request.POST:
        text = request.POST["texto"]

        lines = text.splitlines()

        for line in lines: 
            if line: 
                parts = line.split("#")

                pacientes = Paciente.objects.filter(dni = parts[0])

                if len(pacientes) > 0: 
                    paciente = pacientes[0]

                else:
                    paciente = Paciente()

                    paciente.dni = parts[0]

                paciente.nombre = parts[1]

                paciente.sexo = 1 if parts[3] == "M" else 0 
                
                #edad
                values = parts[2].split("/")

                values = [int(elem) for elem in values]

                paciente.nacimiento = date(values[2], values[0], values[1])
                
                paciente.hospital = Hospital.objects.filter(id = 0)[0]

                paciente.historia = parts[8]
                paciente.direccion = parts[9].capitalize()
                
                paciente.hta = True
                paciente.dm = True
                paciente.fechaHta = date(2021, 1, 1)
                paciente.fechaDm = date(2021, 1, 1)

                paciente.celular = parts[10]

                paciente.save()

        context = {"texto": text}
    else: 
        context = {}

    
    return render(request, "_admin.html", context)
